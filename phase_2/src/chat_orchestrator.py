"""Conversation orchestration and Phase 1 integration for Phase 2."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook

from pathlib import Path
import sys

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from phase_1.src.main import run_phase_1
from .input_parser import (
    detect_missing_fields,
    empty_task,
    merge_task_updates,
    parse_natural_language_request,
)
from .response_formatter import (
    format_missing_prompt,
    format_task_summary,
    format_workflow_result,
)
from .session_manager import SessionManager


@dataclass
class ConversationState:
    session_id: str
    task: dict[str, Any]
    has_rationale: bool = False
    rationale_path: str = ""
    latest_result: dict[str, Any] | None = None


class ChatOrchestrator:
    def __init__(self, session_manager: SessionManager | None = None) -> None:
        self.session_manager = session_manager or SessionManager()

    def init_state(self, existing_session_id: str | None = None) -> ConversationState:
        session_id = self.session_manager.get_or_create_session_id(existing_session_id)
        self.session_manager.ensure_paths(session_id)
        return ConversationState(session_id=session_id, task=empty_task())

    def handle_user_message(self, state: ConversationState, message: str) -> str:
        parsed = parse_natural_language_request(message)
        state.task = merge_task_updates(state.task, parsed)

        missing = detect_missing_fields(state.task)
        summary = format_task_summary(state.task)
        prompt = format_missing_prompt(missing_fields=missing, has_rationale=state.has_rationale)
        return f"{summary}\n\n{prompt}"

    def handle_rationale_upload(
        self,
        state: ConversationState,
        file_name: str,
        file_bytes: bytes,
    ) -> str:
        if not file_name.lower().endswith(".docx"):
            return "Please upload a .docx rationale document."

        saved_path = self.session_manager.save_uploaded_rationale(file_bytes, state.session_id)
        state.has_rationale = True
        state.rationale_path = str(saved_path)

        missing = detect_missing_fields(state.task)
        prompt = format_missing_prompt(missing_fields=missing, has_rationale=state.has_rationale)
        return f"I received the rationale document.\n\n{prompt}"

    def ready_to_run(self, state: ConversationState) -> tuple[bool, str]:
        missing = detect_missing_fields(state.task)
        if missing:
            return False, "Missing required input: parameter_question."
        if not state.has_rationale:
            return False, "Rationale .docx upload is required before running."
        return True, "Ready to run."

    def run_workflow(self, state: ConversationState) -> tuple[str, list[dict[str, str]], Path | None]:
        ready, reason = self.ready_to_run(state)
        if not ready:
            return reason, [], None

        paths = self.session_manager.ensure_paths(state.session_id)
        self.session_manager.save_task_json(state.task, state.session_id)

        result = run_phase_1(
            task_json_path=paths.task_json,
            rationale_docx_path=paths.rationale_docx,
            prep_direction_json_path=None,
            output_xlsx_path=paths.output_xlsx,
        )
        state.latest_result = result

        publication_rows = self._load_publication_rows(paths.output_xlsx)
        return format_workflow_result(result), publication_rows, paths.output_xlsx

    @staticmethod
    def _load_publication_rows(output_xlsx: Path) -> list[dict[str, str]]:
        if not output_xlsx.exists():
            return []
        wb = load_workbook(output_xlsx)
        ws = wb.active
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(
                {
                    "Publication Citation": str(row[0] or ""),
                    "Rationale of Selection": str(row[1] or ""),
                    "Limitations": str(row[2] or ""),
                    "Ranking": str(row[3] or ""),
                }
            )
        return rows
