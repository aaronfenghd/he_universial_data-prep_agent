"""Excel export for selected publications with strict Phase 1 column requirements."""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook, load_workbook

from .publication_selector import ALLOWED_RANKINGS, SelectedPublication

REQUIRED_COLUMNS = [
    "Publication Citation",
    "Rationale of Selection",
    "Limitations",
    "Ranking",
]


def export_to_excel(selected: list[SelectedPublication], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Phase1 Results"

    ws.append(REQUIRED_COLUMNS)

    for item in selected:
        if item.ranking not in ALLOWED_RANKINGS:
            raise ValueError(f"Ranking must be one of {sorted(ALLOWED_RANKINGS)}")
        ws.append(
            [
                item.publication_citation,
                item.rationale_of_selection,
                item.limitations,
                item.ranking,
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def validate_excel(output_path: Path) -> None:
    """Lightweight validation: columns order and ranking values."""
    wb = load_workbook(output_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if header != REQUIRED_COLUMNS:
        raise ValueError(f"Excel columns mismatch. Expected {REQUIRED_COLUMNS}, got {header}")

    for row_idx in range(2, ws.max_row + 1):
        ranking = ws.cell(row=row_idx, column=4).value
        if ranking not in ALLOWED_RANKINGS:
            raise ValueError(f"Invalid ranking in row {row_idx}: {ranking}")
